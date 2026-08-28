# -*- coding: utf-8 -*-
# Предварительно создать БД climate и все таблицы статистики (sql.txt)
from __future__ import annotations
from core.config import STATISTICS_PATH, DB_PARAMS, SECTION_TO_INDUSTRY, PERIOD_PATTERNS, TOP_LEVEL_HINTS

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Iterable

from datetime import date

from services.statistics_periods import (
    period_dates,
)

import openpyxl
import psycopg2

@dataclass
class LeafRecord:
    section_name: str
    industry_name: str
    indicator_name: str
    unit_name: str
    period_label: str
    year: int
    value: float

def is_empty(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")

def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).replace(".0", "")
    text = str(value).replace("\n", " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def is_period_label(text: str) -> bool:
    t = clean_text(text).lower()
    return any(p.search(t) for p in PERIOD_PATTERNS)


def infer_period_type(period_label: str) -> str:
    t = clean_text(period_label).lower()
    if t.startswith("на "):
        return "дата"
    return "период"


def normalize_period_name(period_label: str, year: int) -> str:
    p = clean_text(period_label)
    if p.lower() == "год":
        return f"{year} год"
    return f"{p} {year}"


def row_has_any_value(row: Iterable[object]) -> bool:
    return any(not is_empty(v) for v in row)


def is_top_level_heading(label: str, prev_was_data: bool) -> bool:
    low = label.lower()
    if any(low.startswith(h) for h in TOP_LEVEL_HINTS):
        return True
    if prev_was_data and len(label) > 30:
        return True
    return False


def build_indicator_name(stack: list[str], leaf_label: Optional[str]) -> str:
    parts = [clean_text(x) for x in stack if clean_text(x)]
    if leaf_label and clean_text(leaf_label):
        leaf = clean_text(leaf_label)
        if not parts or parts[-1] != leaf:
            parts.append(leaf)
    return " — ".join(parts)


class DB:
    def __init__(self, dry_run: bool = False):
        self.dry_run = dry_run
        self.conn = None
        self.cache: dict[tuple[str, str], int] = {}

    def connect(self):
        self.conn = psycopg2.connect(
            host=DB_PARAMS["host"],
            port=DB_PARAMS["port"],
            dbname=DB_PARAMS["database"],
            user=DB_PARAMS["user"],
            password=DB_PARAMS["password"],
        )

    def close(self):
        if self.conn is not None:
            self.conn.close()

    def commit(self):
        if self.conn is not None:
            self.conn.commit()

    def rollback(self):
        if self.conn is not None:
            self.conn.rollback()

    def get_or_create_simple(self, table: str, name: str, pk: str) -> int:
        key = (table, name)
        if key in self.cache:
            return self.cache[key]
        if self.dry_run:
            fake_id = len(self.cache) + 1
            self.cache[key] = fake_id
            return fake_id

        sql_select = f"SELECT {pk} FROM {table} WHERE name = %s"
        sql_insert = f"INSERT INTO {table} (name) VALUES (%s) RETURNING {pk}"
        with self.conn.cursor() as cur:
            cur.execute(sql_select, (name,))
            row = cur.fetchone()
            if row:
                obj_id = row[0]
            else:
                cur.execute(sql_insert, (name,))
                obj_id = cur.fetchone()[0]
        self.cache[key] = obj_id
        return obj_id

    def get_or_create_territory(self, territory_name: str, territory_type_name: str = "муниципальный район") -> int:
        tt_id = self.get_or_create_simple("territory_type", territory_type_name, "id")
        key = ("territory", territory_name)
        if key in self.cache:
            return self.cache[key]
        if self.dry_run:
            fake_id = len(self.cache) + 1
            self.cache[key] = fake_id
            return fake_id
        with self.conn.cursor() as cur:
            cur.execute(
                """
                SELECT id
                FROM territory
                WHERE name = %s AND territory_type_id = %s AND parent_territory_id IS NULL
                """,
                (territory_name, tt_id),
            )
            row = cur.fetchone()
            if row:
                territory_id = row[0]
            else:
                cur.execute(
                    """
                    INSERT INTO territory (parent_territory_id, territory_type_id, name)
                    VALUES (NULL, %s, %s)
                    RETURNING id
                    """,
                    (tt_id, territory_name),
                )
                territory_id = cur.fetchone()[0]
        self.cache[key] = territory_id
        return territory_id

    def get_or_create_section(self, industry_name: str, section_name: str) -> int:
        industry_id = self.get_or_create_simple("industry", industry_name, "id")
        key = ("section", section_name)
        if key in self.cache:
            return self.cache[key]
        if self.dry_run:
            fake_id = len(self.cache) + 1
            self.cache[key] = fake_id
            return fake_id
        with self.conn.cursor() as cur:
            cur.execute("SELECT id FROM section WHERE name = %s", (section_name,))
            row = cur.fetchone()
            if row:
                section_id = row[0]
            else:
                cur.execute(
                    "INSERT INTO section (industry_id, name) VALUES (%s, %s) RETURNING id",
                    (industry_id, section_name),
                )
                section_id = cur.fetchone()[0]
        self.cache[key] = section_id
        return section_id

    def get_or_create_indicator(self, section_id: int, unit_id: int, indicator_name: str) -> int:
        key = ("indicator", f"{section_id}::{indicator_name}")
        if key in self.cache:
            return self.cache[key]
        if self.dry_run:
            fake_id = len(self.cache) + 1
            self.cache[key] = fake_id
            return fake_id
        with self.conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM indicator WHERE section_id = %s AND name = %s",
                (section_id, indicator_name),
            )
            row = cur.fetchone()
            if row:
                indicator_id = row[0]
            else:
                cur.execute(
                    """
                    INSERT INTO indicator (section_id, unit_id, name)
                    VALUES (%s, %s, %s)
                    RETURNING id
                    """,
                    (section_id, unit_id, indicator_name),
                )
                indicator_id = cur.fetchone()[0]
        self.cache[key] = indicator_id
        return indicator_id

    def get_or_create_period(
        self,
        period_type_name: str,
        period_name: str,
        start_date: date,
        end_date: date,
        ) -> int:

        period_type_id = (
            self.get_or_create_simple(
                "period_type",
                period_type_name,
                "id",
            )
        )

        key = (
            "period",
            f"{period_type_id}::{period_name}",
        )

        if key in self.cache:
            return self.cache[key]

        if self.dry_run:
            fake_id = len(
                self.cache
            ) + 1

            self.cache[key] = fake_id

            return fake_id

        with self.conn.cursor() as cur:

            cur.execute(
                """
                SELECT
                    id,
                    start_date,
                    end_date
                FROM period
                WHERE period_type_id = %s
                AND name = %s
                """,
                (
                    period_type_id,
                    period_name,
                ),
            )

            row = cur.fetchone()

            if row:
                period_id = row[0]

                current_start = row[1]
                current_end = row[2]

                if (
                    current_start != start_date
                    or current_end != end_date
                ):
                    cur.execute(
                        """
                        UPDATE period
                        SET
                            start_date = %s,
                            end_date = %s
                        WHERE period_id = %s
                        """,
                        (
                            start_date,
                            end_date,
                            period_id,
                        ),
                    )

            else:
                cur.execute(
                    """
                    INSERT INTO period (
                        period_type_id,
                        name,
                        start_date,
                        end_date
                    )
                    VALUES (%s, %s, %s, %s)
                    RETURNING period_id
                    """,
                    (
                        period_type_id,
                        period_name,
                        start_date,
                        end_date,
                    ),
                )

                period_id = (
                    cur.fetchone()[0]
                )

        self.cache[key] = period_id

        return period_id

    def upsert_statistic(self, territory_id: int, indicator_id: int, period_id: int, value: float):
        if self.dry_run:
            return
        with self.conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO statistic (territory_id, indicator_id, period_id, value)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (territory_id, indicator_id, period_id)
                DO UPDATE SET value = EXCLUDED.value
                """,
                (territory_id, indicator_id, period_id, value),
            )


def parse_sheet(ws) -> list[LeafRecord]:
    section_name = clean_text(ws.cell(1, 1).value)
    if not section_name:
        return []
    industry_name = SECTION_TO_INDUSTRY.get(section_name, "экономика")

    years = {}
    for col in range(3, ws.max_column + 1):
        val = ws.cell(2, col).value
        if isinstance(val, int):
            years[col] = val
        elif isinstance(val, float) and float(val).is_integer():
            years[col] = int(val)

    records: list[LeafRecord] = []
    stack: list[str] = []
    prev_was_data = False

    for row in range(3, ws.max_row + 1):
        label = clean_text(ws.cell(row, 1).value)
        unit = clean_text(ws.cell(row, 2).value)
        values = [ws.cell(row, c).value for c in years.keys()]

        if not label and not unit and not row_has_any_value(values):
            continue

        has_numeric = any(isinstance(v, (int, float)) for v in values if v is not None)
        unit_present = bool(unit)

        # Заголовок / группировка без единицы измерения и без значений
        if label and not unit_present and not has_numeric:
            if is_top_level_heading(label, prev_was_data) or not stack:
                stack = [label]
            else:
                # Мягкая иерархия: оставляем корень и меняем/добавляем текущий подуровень
                if len(stack) == 1:
                    stack.append(label)
                else:
                    stack[-1] = label
            prev_was_data = False
            continue

        # Строка-значение
        if unit_present and has_numeric:
            if is_period_label(label):
                indicator_name = build_indicator_name(stack, None)
                period_label = label
            else:
                # самостоятельный показатель или leaf-подпоказатель
                if stack and stack[0] != label:
                    indicator_name = build_indicator_name(stack, label)
                else:
                    indicator_name = clean_text(label)
                period_label = "год"

            # Уточнение имени, если в одной секции одинаковое название встречается с разными единицами
            # Это бывает редко, но защищает от конфликта UNIQUE(section_id, name)
            if section_name == "Предприятия по переработке отходов" and "Количество вывезенных отходов" in indicator_name:
                indicator_name = f"{indicator_name} [{unit}]"

            for col, year in years.items():
                value = ws.cell(row, col).value
                if value is None or value == "":
                    continue
                if not isinstance(value, (int, float)):
                    continue
                records.append(
                    LeafRecord(
                        section_name=section_name,
                        industry_name=industry_name,
                        indicator_name=indicator_name,
                        unit_name=unit,
                        period_label=period_label,
                        year=year,
                        value=float(value),
                    )
                )
            prev_was_data = True
            continue

        # Бывает строка без unit, но со значениями — считаем ее обычным показателем
        if label and has_numeric:
            indicator_name = clean_text(label)
            unit_name = unit if unit else "единица"
            for col, year in years.items():
                value = ws.cell(row, col).value
                if value is None or value == "":
                    continue
                if not isinstance(value, (int, float)):
                    continue
                records.append(
                    LeafRecord(
                        section_name=section_name,
                        industry_name=industry_name,
                        indicator_name=indicator_name,
                        unit_name=unit_name,
                        period_label="год",
                        year=year,
                        value=float(value),
                    )
                )
            prev_was_data = True

    return records


def parse_workbook(xlsx_path: Path) -> list[LeafRecord]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_records: list[LeafRecord] = []
    for ws in wb.worksheets:
        if ws.title == "Оглавление":
            continue
        all_records.extend(parse_sheet(ws))
    return all_records


def load_records(db: DB, records: list[LeafRecord], territory_name: str):
    territory_id = db.get_or_create_territory(territory_name)

    for rec in records:
        unit_id = db.get_or_create_simple("unit", rec.unit_name, "id")
        section_id = db.get_or_create_section(rec.industry_name, rec.section_name)
        indicator_id = db.get_or_create_indicator(section_id, unit_id, rec.indicator_name)
        period_type_name = infer_period_type(
            rec.period_label
        )
        period_name = normalize_period_name(
            rec.period_label,
            rec.year,
        )
        start_date, end_date = period_dates(
            rec.period_label,
            rec.year,
        )
        period_id = db.get_or_create_period(
            period_type_name,
            period_name,
            start_date,
            end_date,
        )
        db.upsert_statistic(territory_id, indicator_id, period_id, rec.value)


def validate_period_dates(
    db: DB,
) -> None:

    with db.conn.cursor() as cur:

        cur.execute(
            """
            SELECT name
            FROM period
            WHERE start_date IS NULL
               OR end_date IS NULL
            ORDER BY name
            LIMIT 20
            """
        )

        rows = cur.fetchall()

    if rows:
        invalid_names = [
            row[0]
            for row in rows
        ]

        raise RuntimeError(
            "Обнаружены периоды "
            "без start_date/end_date: "
            + ", ".join(
                invalid_names
            )
        )

def main():

    db = DB(dry_run=False)

    try:
        print("=== Старт загрузки ===")
        print("Абсолютный путь:", STATISTICS_PATH.resolve())
        print("Содержимое папки:")

        for f in Path(STATISTICS_PATH).iterdir():
            print(" -", f.name)

        print("Подключение к БД...")
        db.connect()
        print("БД подключена")

        files = list(STATISTICS_PATH.glob("*.xlsx"))

        print("Найдено Excel файлов:", len(files))

        if not files:
            print("Excel файлов нет")
            return

        total_records = 0

        for xlsx_path in files:

            print("\n-----------------------")
            print("Файл:", xlsx_path)

            territory_name = xlsx_path.stem

            print("Территория:", territory_name)

            print("Чтение Excel...")
            records = parse_workbook(xlsx_path)

            print("Найдено записей:", len(records))

            if records:
                print("Первая запись:")
                print(records[0])

                print("Секций:",
                      len({r.section_name for r in records}))

                print("Индикаторов:",
                      len({r.indicator_name for r in records}))

                print("Единиц:",
                      len({r.unit_name for r in records}))

                print("Периодов:",
                      len({(r.year, r.period_label) for r in records}))

            else:
                print("Записей нет, пропуск")
                continue

            print("Загрузка в БД...")

            load_records(
                db,
                records,
                territory_name
            )

            print("Загружено:", len(records))

            total_records += len(records)

        print(
            "Проверяем даты периодов..."
        )

        validate_period_dates(
            db
        )

        print(
            "Все периоды имеют даты"
        )

        print("\nСохраняем транзакцию...")
        db.commit()

        print("=== Загрузка завершена ===")
        print("Всего записей:", total_records)

    except Exception as e:
        print("ОШИБКА:")
        print(type(e).__name__)
        print(e)

        db.rollback()
        raise

    finally:
        db.close()
        print("Соединение закрыто")

if __name__ == "__main__":
    main()
